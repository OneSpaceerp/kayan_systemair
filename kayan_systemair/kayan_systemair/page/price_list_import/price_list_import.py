"""
Price List Import Page — Server-Side Methods
=============================================
Provides whitelisted API methods for:
1. Uploading the item-group mapping (article_no → Item Group)
2. Previewing an uploaded price-list Excel file (first 20 rows)
3. Enqueueing the background import job
4. Getting import progress / log

Import file column contract (§5 of QUOTATION_ITEM_SPEC):
  Item no | Item name | Sales price | Currency | Item Group | Type of Fan |
  Family name | Product group descr | Business area descr |
  Available from Primary Factory | Temperature Rate of Motor | …
"""

import frappe
from frappe import _
from frappe.utils import flt, nowdate


# ---------------------------------------------------------------------------
# Step 0 — Item Group Mapping
# ---------------------------------------------------------------------------

@frappe.whitelist()
def upload_item_group_mapping(file_url):
    """
    Parse the Item Group mapping Excel file and bulk-upsert every row into
    the SystemAir Item Group Map doctype.

    Performance: uses chunked INSERT ... ON DUPLICATE KEY UPDATE instead of
    per-row frappe.db.exists()/insert() — a 17k-row file previously issued
    ~50k queries and hit the gateway 504 timeout; this version completes in
    a few seconds inside the same request (no JS changes required).

    Expected columns (case-insensitive): "Item no" and "Item Group".
    The header row is searched within the first 5 rows, so both the generated
    mapping files (header on row 1) and the client's original Item Group.xlsx
    (blank rows first, header on row 3) work.

    Missing Item Groups are auto-created under the 'SystemAir Fans' parent
    group instead of causing rows to be skipped.

    Returns:
        dict: {"loaded": int, "skipped": int, "groups_created": int}
    """
    import openpyxl

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        frappe.throw(_("Mapping file is empty."))

    # ------------------------------------------------------------------ #
    # Detect the header row (scan first 5 rows) and column positions      #
    # ------------------------------------------------------------------ #
    article_col = group_col = header_idx = None
    for r_idx, r in enumerate(rows[:5]):
        header = [str(c).strip().lower() if c is not None else "" for c in r]
        a = g = None
        for i, h in enumerate(header):
            if a is None and any(k in h for k in ["item no", "item_no", "article no", "article_no", "artikelnr"]):
                a = i
            if g is None and h in ("item group", "item_group", "itemgroup"):
                g = i
        if a is not None and g is not None:
            article_col, group_col, header_idx = a, g, r_idx
            break

    if article_col is None or group_col is None:
        detected = ", ".join(
            str(c).strip() if c is not None else "" for c in rows[0]
        )
        frappe.throw(
            _(
                "Could not find 'Item no' and 'Item Group' columns in the "
                "first 5 rows of the mapping file. First row: {0}"
            ).format(detected)
        )

    # ------------------------------------------------------------------ #
    # Collect and dedupe pairs (last occurrence wins)                     #
    # ------------------------------------------------------------------ #
    pairs = {}
    skipped = 0
    for row in rows[header_idx + 1:]:
        if not row:
            skipped += 1
            continue
        article_no = (
            str(row[article_col]).strip()
            if article_col < len(row) and row[article_col] is not None else ""
        )
        item_group = (
            str(row[group_col]).strip()
            if group_col < len(row) and row[group_col] is not None else ""
        )
        if not article_no or not item_group:
            skipped += 1
            continue
        pairs[article_no] = item_group

    if not pairs:
        frappe.throw(_("No valid mapping rows found in the file."))

    # ------------------------------------------------------------------ #
    # Auto-create missing Item Groups (small set — one exists-check each) #
    # ------------------------------------------------------------------ #
    groups_created = 0
    existing_groups = set(
        r[0] for r in frappe.db.sql("SELECT name FROM `tabItem Group`")
    )
    for group in set(pairs.values()):
        if group not in existing_groups:
            _ensure_parent_group()
            frappe.get_doc({
                "doctype": "Item Group",
                "item_group_name": group,
                "parent_item_group": "SystemAir Fans",
                "is_group": 0,
            }).insert(ignore_permissions=True)
            groups_created += 1

    # ------------------------------------------------------------------ #
    # Bulk upsert in chunks (name == article_no per doctype autoname)     #
    # ------------------------------------------------------------------ #
    from frappe.utils import now

    ts = now()
    user = frappe.session.user
    items = list(pairs.items())
    chunk_size = 2000
    for start in range(0, len(items), chunk_size):
        chunk = items[start:start + chunk_size]
        values_sql = ", ".join(["(%s, %s, %s, %s, %s, %s, %s, 0, 0)"] * len(chunk))
        params = []
        for article_no, item_group in chunk:
            params.extend([article_no, article_no, item_group, ts, ts, user, user])
        frappe.db.sql(
            f"""
            INSERT INTO `tabSystemAir Item Group Map`
                (name, article_no, item_group, creation, modified,
                 owner, modified_by, docstatus, idx)
            VALUES {values_sql}
            ON DUPLICATE KEY UPDATE
                item_group = VALUES(item_group),
                modified = VALUES(modified),
                modified_by = VALUES(modified_by)
            """,
            params,
        )

    frappe.db.commit()
    return {"loaded": len(pairs), "skipped": skipped, "groups_created": groups_created}


@frappe.whitelist()
def get_mapping_status():
    """Return the number of item-group mapping records currently loaded."""
    count = frappe.db.count("SystemAir Item Group Map")
    return {"count": count}


# ---------------------------------------------------------------------------
# Steps 1 & 2 — Price list selection and file preview
# ---------------------------------------------------------------------------

@frappe.whitelist()
def preview_excel(file_url, price_list):
    """
    Read the first 20 data rows from an uploaded Excel file and return
    them as a list of dicts for preview.  Each row includes the resolved
    Item Group from the mapping or the file's own Item Group column.

    Returns:
        dict: {
            "columns": [...],
            "rows": [...],
            "sheet_name": str,
            "total_rows": int
        }
    """
    import openpyxl

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    sheet_name = _detect_sheet(wb.sheetnames, price_list)
    ws = wb[sheet_name]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        frappe.throw(_("Excel file appears to be empty."))

    header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
    col_map = _map_columns(header_row)

    if not col_map.get("item_name") and not col_map.get("article_no"):
        frappe.throw(
            _(
                "Could not detect required columns. "
                "Expected: 'Item name' or 'item_name', 'Sales price' or 'sales_price', "
                "'Item no' or 'item_no'."
            )
        )

    preview_rows = []
    for row in rows[1:21]:
        mapped = _map_row(row, col_map)
        if mapped:
            article_no = mapped.get("article_no") or ""
            # Resolve group: mapping first, file column fallback
            item_group = ""
            if article_no:
                item_group = frappe.db.get_value(
                    "SystemAir Item Group Map", article_no, "item_group"
                ) or ""
            if not item_group:
                item_group = mapped.get("item_group_file") or ""
            mapped["item_group"] = item_group
            preview_rows.append(mapped)

    wb.close()
    return {
        "columns": ["Item Code / Name", "Price (EUR)", "Article No.", "Item Group"],
        "rows": preview_rows,
        "sheet_name": sheet_name,
        "total_rows": max(0, ws.max_row - 1),
    }


@frappe.whitelist()
def start_import(file_url, price_list, sheet_name=None):
    """
    Enqueue a background job to import all rows from the Excel file
    into ERPNext Item Price records.

    If no item-group mapping is loaded AND the price list file itself lacks
    an Item Group column, rows without a resolvable group are skipped.

    Returns:
        dict: {"log_name": str}
    """
    log = frappe.get_doc({
        "doctype": "SystemAir Import Log",
        "price_list": price_list,
        "sheet_name": sheet_name or "",
        "imported_by": frappe.session.user,
        "status": "Pending",
        "records_created": 0,
        "records_updated": 0,
        "records_skipped": 0,
    })
    log.insert(ignore_permissions=True)
    frappe.db.commit()

    frappe.enqueue(
        "kayan_systemair.kayan_systemair.page.price_list_import.price_list_import._run_import",
        queue="long",
        timeout=3600,
        file_url=file_url,
        price_list=price_list,
        sheet_name=sheet_name,
        log_name=log.name,
    )

    return {"log_name": log.name}


@frappe.whitelist()
def get_import_status(log_name):
    """Return the current status of an import job."""
    log = frappe.db.get_value(
        "SystemAir Import Log",
        log_name,
        ["status", "records_created", "records_updated", "records_skipped", "skip_reason"],
        as_dict=True,
    )
    return log or {}


@frappe.whitelist()
def get_price_lists():
    """Return available SystemAir price lists."""
    return frappe.db.get_all(
        "Price List",
        filters={"price_list_name": ["like", "Systemair%"], "enabled": 1},
        fields=["name", "price_list_name", "currency"],
    )


# ---------------------------------------------------------------------------
# Background job (called via frappe.enqueue — not whitelisted)
# ---------------------------------------------------------------------------

def _run_import(file_url, price_list, sheet_name, log_name):
    """
    Background job: read Excel file row by row, create/update Item + Item Price.
    Rows without a resolved item group (from mapping or file column) are skipped.
    New groups found in the file are auto-created under 'SystemAir Fans'.
    Updates the SystemAir Import Log on completion.
    """
    import openpyxl

    created = 0
    updated = 0
    skipped = 0
    skip_reasons = []
    error_msg = None
    _created_groups = set()  # cache groups created in this import run

    try:
        file_doc = frappe.get_doc("File", {"file_url": file_url})
        file_path = file_doc.get_full_path()

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if not sheet_name:
            sheet_name = _detect_sheet(wb.sheetnames, price_list)
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel file is empty")

        header_row = [str(c).strip() if c is not None else "" for c in rows[0]]
        col_map = _map_columns(header_row)

        if not col_map.get("item_name") and not col_map.get("article_no"):
            raise ValueError(
                "Cannot detect required columns (item_name/item_no, sales_price)"
            )

        for row_data in rows[1:]:
            mapped = _map_row(row_data, col_map)
            if not mapped:
                skipped += 1
                continue

            item_name  = mapped.get("item_name") or mapped.get("article_no")
            article_no = mapped.get("article_no") or ""
            price      = flt(mapped.get("price"))

            if not item_name or not price:
                skipped += 1
                continue

            # --- Resolve item group (mapping first, file column fallback) ---
            item_group = None
            if article_no:
                item_group = frappe.db.get_value(
                    "SystemAir Item Group Map", article_no, "item_group"
                )
            if not item_group:
                item_group_file = mapped.get("item_group_file") or ""
                if item_group_file:
                    # Auto-create group under SystemAir Fans if not present
                    if item_group_file not in _created_groups:
                        if not frappe.db.exists("Item Group", item_group_file):
                            _ensure_parent_group()
                            frappe.get_doc({
                                "doctype": "Item Group",
                                "item_group_name": item_group_file,
                                "parent_item_group": "SystemAir Fans",
                                "is_group": 0,
                            }).insert(ignore_permissions=True)
                        _created_groups.add(item_group_file)
                    item_group = item_group_file

            if not item_group:
                skipped += 1
                if len(skip_reasons) < 20:
                    skip_reasons.append(
                        f"{article_no or item_name}: no item group mapping found"
                    )
                continue

            # --- Get or create Item with enriched attributes ---
            item_code = _get_or_create_item(
                item_name,
                article_no=article_no,
                item_group=item_group,
                type_of_fan=mapped.get("type_of_fan") or "",
                family_name=mapped.get("family_name") or "",
                primary_factory=mapped.get("primary_factory") or "",
                temperature_rate=mapped.get("temperature_rate") or "",
            )

            was_created = _upsert_item_price(item_code, price_list, price)
            if was_created:
                created += 1
            else:
                updated += 1

        wb.close()
        frappe.db.commit()

    except Exception as e:
        error_msg = str(e)
        frappe.log_error(frappe.get_traceback(), "SystemAir Price List Import Error")

    status = "Failed" if error_msg else "Completed"
    skip_reason_text = "\n".join(skip_reasons) if skip_reasons else ""
    if error_msg:
        skip_reason_text = (error_msg + "\n" + skip_reason_text).strip()

    frappe.db.set_value(
        "SystemAir Import Log",
        log_name,
        {
            "status": status,
            "records_created": created,
            "records_updated": updated,
            "records_skipped": skipped,
            "skip_reason": skip_reason_text[:2000] if skip_reason_text else "",
        },
    )
    frappe.db.commit()


# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def _ensure_parent_group():
    """Guarantee the 'SystemAir Fans' parent group exists."""
    if frappe.db.exists("Item Group", "SystemAir Fans"):
        return
    parent = "Products" if frappe.db.exists("Item Group", "Products") else "All Item Groups"
    frappe.get_doc({
        "doctype": "Item Group",
        "item_group_name": "SystemAir Fans",
        "parent_item_group": parent,
        "is_group": 1,
    }).insert(ignore_permissions=True)


def _detect_sheet(sheet_names, price_list):
    """Auto-detect the correct sheet name based on the price list."""
    price_list_lower = price_list.lower()
    for name in sheet_names:
        name_lower = name.lower()
        if "germany" in price_list_lower and ("germany" in name_lower or "de" in name_lower):
            return name
        if "malaysia" in price_list_lower and ("malaysia" in name_lower or "my" in name_lower):
            return name
    return sheet_names[0]


def _map_columns(header_row):
    """
    Build a column-index map from the header row.
    Handles Germany / Malaysia enriched price list columns (§5) plus German headers.

    Returns dict with any subset of:
      item_name, price, article_no, item_group_file,
      type_of_fan, family_name, primary_factory, temperature_rate
    """
    mapping = {}
    name_candidates     = ["item name", "item_name", "description", "bezeichnung", "artikel"]
    price_candidates    = ["sales price", "sales_price", "price", "preis", "list price"]
    article_candidates  = ["item no", "item_no", "article no", "article_no", "artikel nr", "artikelnr"]
    # "item group" must be an exact / near-exact match to avoid colliding with
    # "product group descr" or "business area descr" columns.
    group_candidates    = ["item group", "item_group", "itemgroup"]
    fan_type_candidates = ["type of fan", "type_of_fan"]
    family_candidates   = ["family name", "family_name"]
    factory_candidates  = ["available from primary factory", "primary factory", "primary_factory"]
    temp_candidates     = ["temperature rate of motor", "temperature rate", "temperature_rate"]

    for idx, col in enumerate(header_row):
        lc = col.lower().strip()
        if "item_name"       not in mapping and any(c in lc for c in name_candidates):
            mapping["item_name"] = idx
        if "price"           not in mapping and any(c in lc for c in price_candidates):
            mapping["price"] = idx
        if "article_no"      not in mapping and any(c in lc for c in article_candidates):
            mapping["article_no"] = idx
        if "item_group_file" not in mapping and lc in group_candidates:
            mapping["item_group_file"] = idx
        if "type_of_fan"     not in mapping and any(c in lc for c in fan_type_candidates):
            mapping["type_of_fan"] = idx
        if "family_name"     not in mapping and any(c in lc for c in family_candidates):
            mapping["family_name"] = idx
        if "primary_factory" not in mapping and any(c in lc for c in factory_candidates):
            mapping["primary_factory"] = idx
        if "temperature_rate" not in mapping and any(c in lc for c in temp_candidates):
            mapping["temperature_rate"] = idx

    return mapping


def _map_row(row, col_map):
    """Map a data row tuple to a dict using col_map."""
    if not row:
        return None

    def safe_get(key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    item_name  = safe_get("item_name")
    price      = safe_get("price")
    article_no = safe_get("article_no")

    if not item_name and not article_no:
        return None

    def s(val):
        return str(val).strip() if val is not None else ""

    return {
        "item_name":       s(item_name),
        "price":           flt(price) if price is not None else 0.0,
        "article_no":      s(article_no),
        "item_group_file": s(safe_get("item_group_file")),
        "type_of_fan":     s(safe_get("type_of_fan")),
        "family_name":     s(safe_get("family_name")),
        "primary_factory": s(safe_get("primary_factory")),
        "temperature_rate": s(safe_get("temperature_rate")),
    }


# Normalise Type of Fan values from the price list to valid Select options
_TYPE_OF_FAN_MAP = {
    "centrifugal fan":      "Centrifugal Fan",
    "centrifugal roof top": "Centrifugal Roof Top",
    "inline fan":           "Inline Fan",
    "axial inline":         "Axial Inline",
    "wall mounted":         "Wall Mounted",
    "induction jet fan":    "Induction Jet Fan",
    "impulse jet fan":      "Impulse Jet Fan",
    "accessories":          "Accessories",
    "centrifugal box fan":  "Centrifugal Box Fan",
}

_TEMPERATURE_RATE_OPTIONS = {
    "120°c continuous", "300°c/2hr", "400°c/2hr", "600°c/2hr", "explosion proof"
}


def _normalize_type_of_fan(val):
    return _TYPE_OF_FAN_MAP.get(val.lower().strip(), "") if val else ""


def _normalize_temperature_rate(val):
    """Return val if it matches a valid option (case-insensitive), else ''."""
    if not val:
        return ""
    lc = val.lower().strip()
    for opt in _TEMPERATURE_RATE_OPTIONS:
        if opt == lc:
            # Return the canonical casing stored in the Select options
            canon = {
                "120°c continuous": "120°C continuous",
                "300°c/2hr": "300°C/2hr",
                "400°c/2hr": "400°C/2hr",
                "600°c/2hr": "600°C/2hr",
                "explosion proof": "Explosion proof",
            }
            return canon.get(lc, val)
    return ""


def _get_or_create_item(
    item_name,
    article_no="",
    item_group="SystemAir Axial Fans",
    type_of_fan="",
    family_name="",
    primary_factory="",
    temperature_rate="",
):
    """
    Return existing item_code or create a new Item and return its code.
    Always updates enrichment fields (article_no, type, family, factory,
    temperature_rate) on both create and update paths.
    """
    existing = frappe.db.get_value("Item", {"item_code": item_name}, "name")
    if not existing:
        existing = frappe.db.get_value("Item", {"item_name": item_name}, "name")

    clean_fan_type   = _normalize_type_of_fan(type_of_fan)
    clean_temp_rate  = _normalize_temperature_rate(temperature_rate)

    if existing:
        update_data = {}
        if article_no:
            update_data["sa_article_no"]      = article_no[:140]
        if clean_fan_type:
            update_data["sa_type_of_fan"]     = clean_fan_type
        if family_name:
            update_data["sa_product_family"]  = family_name[:140]
        if primary_factory:
            update_data["sa_primary_factory"] = primary_factory[:140]
        if clean_temp_rate:
            update_data["sa_temperature_rate"] = clean_temp_rate
        if update_data:
            frappe.db.set_value("Item", existing, update_data)
        return existing

    item = frappe.get_doc({
        "doctype": "Item",
        "item_code": item_name[:140],
        "item_name": item_name[:140],
        "item_group": item_group,
        "stock_uom": "Nos",
        "is_purchase_item": 1,
        "is_sales_item": 1,
        "is_stock_item": 0,
        "sa_article_no":      article_no[:140] if article_no else "",
        "sa_type_of_fan":     clean_fan_type,
        "sa_product_family":  family_name[:140] if family_name else "",
        "sa_primary_factory": primary_factory[:140] if primary_factory else "",
        "sa_temperature_rate": clean_temp_rate,
    })
    item.insert(ignore_permissions=True)
    return item.item_code


def _upsert_item_price(item_code, price_list, price):
    """
    Create or update an Item Price record.

    Returns:
        bool: True if created, False if updated.
    """
    existing = frappe.db.get_value(
        "Item Price",
        {"item_code": item_code, "price_list": price_list, "selling": 1},
        "name",
    )

    if existing:
        frappe.db.set_value("Item Price", existing, "price_list_rate", price)
        return False
    else:
        frappe.get_doc({
            "doctype": "Item Price",
            "item_code": item_code,
            "price_list": price_list,
            "price_list_rate": price,
            "selling": 1,
            "currency": "EUR",
            "valid_from": nowdate(),
        }).insert(ignore_permissions=True)
        return True
